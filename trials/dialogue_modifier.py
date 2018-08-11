import nltk
import os
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords
from nltk.corpus import wordnet, brown
from openpyxl import load_workbook
import logging
import sqlite3
freq_conn = sqlite3.connect("/Users/karsomas/BITS/Project/data/sqlite_dbs/WordFrequency.db")

freq_cursor = freq_conn.cursor()

'''
from https://pythonprogramming.net/part-of-speech-tagging-nltk-tutorial/
POS tag list:

CC	coordinating conjunction
CD	cardinal digit
DT	determiner
EX	existential there (like: "there is" ... think of it like "there exists")
FW	foreign word
IN	preposition/subordinating conjunction
JJ	adjective	'big'
JJR	adjective, comparative	'bigger'
JJS	adjective, superlative	'biggest'
LS	list marker	1)
MD	modal	could, will
NN	noun, singular 'desk'
NNS	noun plural	'desks'
NNP	proper noun, singular	'Harrison'
NNPS	proper noun, plural	'Americans'
PDT	predeterminer	'all the kids'
POS	possessive ending	parent\'s
PRP	personal pronoun	I, he, she
PRP$	possessive pronoun	my, his, hers
RB	adverb	very, silently,
RBR	adverb, comparative	better
RBS	adverb, superlative	best
RP	particle	give up
TO	to	go 'to' the store.
UH	interjection	errrrrrrrm
VB	verb, base form	take
VBD	verb, past tense	took
VBG	verb, gerund/present participle	taking
VBN	verb, past participle	taken
VBP	verb, sing. present, non-3d	take
VBZ	verb, 3rd person sing. present	takes
WDT	wh-determiner	which
WP	wh-pronoun	who, what
WP$	possessive wh-pronoun	whose
WRB	wh-abverb	where, when
'''

filt_pos = ["VB", "VBD", "VBG", "VBN", "VBP", "VBZ", "NN", "NNS", "NNP", "NNPS", "JJ", "JJR", "JJS", "RB", "RBR", "RBS"]
univ_pos_map = {"v": ["VERB"],
           "n": ["NOUN"],
           "a": ["ADJ"],
           "r": ["ADV"]}
nltk_pos_map = {"v": ["VB", "VBD", "VBG", "VBN", "VBP", "VBZ", "VERB"],
           "n": ["NN", "NNS", "NNP", "NNPS", "NOUN"],
           "a": ["JJ", "JJR", "JJS", "ADJ"],
           "r": ["RB", "RBR", "RBS", "ADV"]}

def mark_pos_for_line(line, category='news'):
    tagged_words = []
    brown_tagged_sents = brown.tagged_sents(categories=category, tagset='universal')
    t0 = nltk.DefaultTagger('NOUN')
    t1 = nltk.UnigramTagger(brown_tagged_sents, backoff=t0)
    t2 = nltk.BigramTagger(brown_tagged_sents, backoff=t1)
    tagged_words = t2.tag(nltk.word_tokenize(line))
    return tagged_words


def get_transformed_line(line, sub_word_map):
    mod_line = ""
    punctuations = '''!()-[]{};:'"\,<>./?@#$%^&*_~'''
    sub_word_list = [w for w in sub_word_map]
    #print "sub_word_list:", sub_word_list
    act_word_list = word_tokenize(line)
    #print "act word list:", act_word_list
    for words in act_word_list:
        if words in sub_word_list:
            mod_line = mod_line + " " + sub_word_map[words]
        elif words in punctuations:
            mod_line = mod_line + words
        else:
            mod_line = mod_line + " " + words
    mod_line +="\n"
    #print "modified line:", mod_line
    return mod_line.strip(" ")


def get_best_match(word_list, user_score=0):
    #print "received word list:", word_list
    for keys in word_list:
        #print "For Key:", keys, ", No. of suggestions:", len(word_list[keys]), " Suggestions:", word_list[keys]
        if len(word_list[keys]) > 0:
            #print "for word", keys, "selected Suggestion:", word_list[keys][0]
            word_list[keys] = get_best_fit(keys, word_list[keys])
        else:
            #print "Selecting the same word"
            word_list[keys] = keys
    return word_list


def get_mapped_pos(pos, tag_model='nltk'):
    if tag_model == 'nltk':
        for key in nltk_pos_map:
            if pos in nltk_pos_map[key]:
                return key
        return ""
    elif tag_model == 'universal':
        for key in univ_pos_map:
            if pos in univ_pos_map[key]:
                return key
        return ""

def get_synonym_list(word, pos=None):
    #logging.debug("Received Word:", word, " PoS:", pos)
    mapped_pos = get_mapped_pos(pos, "universal")
    #print "Received Word:", word, " PoS:", pos, " mapped_pos:", mapped_pos
    sub_list = []
    if mapped_pos != "":
        synsets = wordnet.synsets(word, pos=mapped_pos)
    else:
        synsets = wordnet.synsets(word)
    for ss in synsets:
        sub_list = sub_list + ss.lemma_names()
        for simWords in ss.similar_tos():
            sub_list = sub_list + simWords.lemma_names()

    #print "Alternates:", sub_list
    return sub_list

def filter_words(words):
    stop_words = set(stopwords.words('english'))
    punctuations = '''!()-[]{};:'"\,<>./?@#$%^&*_~'''

    filtered_words = [w for w in words if not w[0] in stop_words and not w[0] in punctuations and len(w[0]) > 2 and w[1] in [pos for pos_list in nltk_pos_map.values() for pos in pos_list]]
    return filtered_words


def extract_dialogue(file_name):

    mod_filename = "mod_" + os.path.basename(file_name)
    write_file_ptr = open(mod_filename, "w")
    file_ptr = open(file_name, "r")
    for line in file_ptr:
        if line[0].isdigit() or line[0].isspace():
            write_file_ptr.write(line)
            continue
        else:
            word_map = {}
            #logging.debug(line)
            print "----------------------------------------"
            print "Line:", line
            tagged = mark_pos_for_line(line)
            # words = word_tokenize(line)
            # tagged = nltk.pos_tag(words)
            #print tagged
            #logging.info("Tagged Words" + tagged)
            filtered_words = filter_words(tagged)
            words = line.split(" ")
            #print "words", words
            #print "filtered words", filtered_words
            #logging.debug("Filtered words:", filtered_words)
            for tup in filtered_words:
                word_map[tup[0]] = get_synonym_list(tup[0], tup[1])
            word_map = get_best_match(word_map)
            #print "Line:", line
            #print "Subs:", word_map
            mod_line = get_transformed_line(line, word_map)
            print "Modified line:", mod_line
            write_file_ptr.write(mod_line)

    write_file_ptr.close()
    file_ptr.close()

    print "fileName:", os.path.basename(file_name)
    print "Modified FileName:", mod_filename


def get_freq_details(words):
    word_list = list(words)
    print "Number of words:", len(word_list)

    formed_clause = ""
    for word in word_list:
        if formed_clause == "":
            formed_clause = "Word=\"" + word +"\""
        else:
            formed_clause = formed_clause + " OR Word=\"" + word + "\""

    formed_cmd = "select Word, SUBTLWF from SUBTLEX_US where " + formed_clause
    results = freq_cursor.execute(formed_cmd).fetchall()
    result_json = {}
    for rows in results:
        result_json[rows[0]] = float(rows[1])
        word_list.remove(rows[0])
    for word in word_list:
        result_json[word] = 0

    return result_json


# def get_freq_info(word):
#     wb2 = load_workbook("/Users/karsomas/BITS/Project/data/word_freq/SUBTLEX-US_frequency_list_with_PoS_information.xlsx", data_only=True)
#     #print wb2.sheetnames[0]
#     ws = wb2[(wb2.sheetnames[0])]
#     row_int = 0
#     found_flag = False
#     for row in ws[('A{}:A{}'.format(ws.min_row, ws.max_row))]:
#         row_int += 1
#         for cell in row:
#             if cell.value == word:
#                 found_flag = True
#                 print "found word:" + word
#                 break
#
#         if found_flag:
#             break
#
#     if found_flag:
#         return ws["F{}".format(row_int)].value
#     else:
#         return -1


def get_best_fit(word, syn_list):
    high_score = 0
    curr_score = 0
    best_word = word
    if len(syn_list) <= 0:
        return word
    else:
        res = get_freq_details(syn_list)
        for key in res:
            curr_score = res[key]
            if curr_score > high_score:
                high_score = curr_score
                best_word = key
    return best_word



#logging.basicConfig(format='%(asctime)s %(levelname)s:%(message)s', level=logging.DEBUG)

#extract_dialogue("/Users/karsomas/BITS/Project/Subtitles/Pirates_of_the_Caribbean_The_Curse_of_the_Black_Pearl.DVDRip.aXXo.en.srt")
extract_dialogue("/Users/karsomas/BITS/Project/Subtitles/Training_subs.srt")
