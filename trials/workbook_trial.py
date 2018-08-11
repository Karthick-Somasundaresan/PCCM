from openpyxl import load_workbook


def get_freq_info(word):
    wb2 = load_workbook("/Users/karsomas/BITS/Project/data/word_freq/sample_word_freq.xlsx", data_only=True)
    print wb2.sheetnames[0]
    ws = wb2[(wb2.sheetnames[0])]
    row_int = 0
    found_flag = False
    for row in ws[('A{}:A{}'.format(ws.min_row, ws.max_row))]:
        row_int += 1
        for cell in row:
            print cell.value
            if cell.value == word:
                found_flag = True
                break

        if found_flag:
            break

    if found_flag:
        return ws["F{}".format(row_int)].value
    else:
        return -1


score = get_freq_info("abatement")

print score




